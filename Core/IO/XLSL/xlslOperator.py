#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2017/2/14
# @Company : INVINCIBLE STUDIO
# @Author  : Mo Wenlong
# @Email   : invincible0918@126.com
# @File    : xlslOperator.py


import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook import defined_name
from openpyxl.worksheet.table import Table


class Sheet(object):
    def __init__(self, workBook, sheetName, isFlat=False):
        """
        :param workBook:
        :param sheetName:
        :param isFlat: make return data to be flatten
        """
        self.__wb = workBook
        self.__sheetName = sheetName

        self.__sheet = self.__wb.get_sheet_by_name(self.__sheetName)

        # key is title by column, value is list
        self.__data = {}

        # first dimension is row, second dimension is column
        def func(cell):
            return str(cell.value) if cell.value != None else ''
        self.__rawData = [[func(cell) for cell in row] for row in self.__sheet.iter_rows()]

        self.__keys = []

        for row in self.__sheet.rows:
            cells = [cell for cell in row]
            if not self.__keys:
                nullCount = 0
                for i in range(len(cells)):
                    if cells[i].value:
                        self.__keys.append(cells[i].value)
                    else:
                        self.__keys.append('null')
                        nullCount += 1

                # need to filter the null cell
                if nullCount == len(cells):
                    self.__keys = []
                else:
                    for key in self.__keys:
                        self.__data[key] = []
            else:
                cellValues = [cell.value for cell in cells]
                if cellValues.count(None) == len(cellValues):
                    break
                else:
                    for i in range(len(cellValues)):
                        self.__data[self.__keys[i]].append(cellValues[i])

    def __getitem__(self, item):
        return self.__data.get(item)

    @property
    def sheet(self):
        return self.__sheet

    @property
    def sheetName(self):
        return self.__sheetName

    @property
    def keys(self):
        return self.__keys

    @property
    def data(self):
        return self.__data

    @property
    def rawData(self):
        return self.__rawData

    def valuesByKey(self, key):
        return self.__data.get(key)


class XlslOperator(object):
    def __init__(self, xlslPath):
        self._xlslPath = xlslPath

        # key is sheet name, value is sheet
        self._data = {}

        wb = openpyxl.load_workbook(filename=self._xlslPath)
        self.__sheetNames = wb.get_sheet_names()
        for sheetName in self.__sheetNames:
            self._data.update({sheetName: Sheet(wb, sheetName)})

    def getSheetByName(self, sheetName):
        return self._data.get(sheetName, None)

    def getAllSheets(self):
        return self._data.values()

    @property
    def sheetNames(self):
        return self.__sheetNames


class XlsxWritter(object):
    def __init__(self):
        pass

    def createBy2DArray(self, sheetName, datas, outputPath):
        """
        :param datas: a 2D array, first dimension is row, second dimension is column

        """
        wb = openpyxl.Workbook()

        # create a new sheet with sheet name, so it's the first sheet
        sheet = wb.create_sheet(sheetName, 0)

        # remove the empty sheet which is after the created one
        wb.remove(wb.worksheets[1])

        for data in datas:
            sheet.append(data)

        wb.save(outputPath)

    def replaceSheet(self, sheet, datas, outputPath):
        """
        :param sheet: original sheet
        :param sheet: a 2D array, first dimension is row, second dimension is column

        """
        wb = openpyxl.load_workbook(outputPath)
        _sheet = wb[sheet.sheetName]

        for i in range(len(datas)):
            for j in range(len(datas[i])):
                _sheet.cell(row=i+1, column=j+1).value = datas[i][j]

        ref = _sheet._tables[0].ref
        tmp = ref.split(':')
        _ref = '%s:%s%s' % (tmp[0], ''.join([c for c in tmp[1] if c.isalpha()]), len(datas)+1)
        _sheet._tables[0].ref = _ref

        wb.save(outputPath)

    def appendSheet(self, sheetName, datas, outputPath):
        """
        :param sheet: original sheet
        :param sheet: a 2D array, first dimension is row, second dimension is column

        """
        wb = openpyxl.load_workbook(outputPath)

        sheet = wb[sheetName]

        # remove the previous sheet
        wb.remove(wb[sheet.sheetName])

        # create a new sheet with sheet name, so it's the first sheet
        _sheet = wb.create_sheet(sheet.sheetName, 0)

        for data in datas:
            _sheet.append(data)

        ranges = sheet.cellRanges
        _ranges = [ranges[0], ranges[1], get_column_letter(len(datas[0])), len(datas)+1]
        # defn = defined_name.DefinedName(sheet.sheetName)
        # defn.value = '$%s$%s:$%s$%s' % (_ranges[0], _ranges[1], _ranges[2], _ranges[3])
        # wb.create_named_range(sheet.sheetName, _sheet, )

        defn = wb.defined_names.get(sheet.sheetName)
        print defn
        defn.value = '$%s$%s:$%s$%s' % (_ranges[0], _ranges[1], _ranges[2], _ranges[3])
        print defn

        # wb.defined_names.append(defn)

        # wb.save(outputPath)
